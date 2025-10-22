from io import BytesIO







from pathlib import Path







from datetime import datetime







from typing import List















import numpy as np







import pandas as pd







import streamlit as st















from core import db, utils
from core.data_table import render_table
from core.theme import load_theme















Alignment = None
Workbook = None
get_column_letter = None







try:







    import openpyxl  # noqa: F401







    from openpyxl.styles import Alignment as _Alignment







    Alignment = _Alignment







    from openpyxl import Workbook as _Workbook
    from openpyxl.utils import get_column_letter as _get_column_letter
    Workbook = _Workbook
    get_column_letter = _get_column_letter
    HAS_OPENPYXL = True







except ModuleNotFoundError:







    HAS_OPENPYXL = False























st.set_page_config(page_title="Fase 0 - Portafolio", page_icon="🌲", layout="wide")

load_theme()















CSS = """
<style>
body { background: linear-gradient(180deg, var(--linen-100) 0%, #f1eadf 60%, #e9e0d2 100%); color: var(--text-900); }
h1, h2, h3 { font-weight: 700; letter-spacing: 0.25px; }

.section-card {
    background: #ffffff;
    border-radius: 22px;
    border: 1px solid rgba(var(--shadow-color), 0.12);
    padding: 1.6rem 1.9rem;
    box-shadow: 0 24px 46px rgba(var(--shadow-color), 0.16);
    margin-bottom: 1.8rem;
}

.badge {
    display: inline-flex;
    align-items: center;
    gap: 0.45rem;
    padding: 6px 16px;
    border-radius: 999px;
    background: rgba(var(--forest-500), 0.12);
    color: var(--forest-700);
    font-weight: 600;
    font-size: 0.85rem;
    border: 1px solid rgba(var(--forest-500), 0.25);
}

.primary-btn button {
    background: linear-gradient(140deg, var(--wood-600), var(--forest-700));
    border: 1px solid rgba(var(--shadow-color), 0.35);
    color: #fefcf8;
    font-weight: 600;
    border-radius: 999px;
    padding: 0.55rem 1.4rem;
    box-shadow: 0 18px 28px rgba(var(--shadow-color), 0.22);
}

.data-editor .stDataFrame {
    border-radius: 18px;
    border: 1px solid rgba(var(--shadow-color), 0.08);
    box-shadow: 0 16px 34px rgba(var(--shadow-color), 0.16);
}

.metric-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(220px, 1fr));
    gap: 1.3rem;
    margin: 1.2rem 0 1.9rem;
}

.metric-card {
    background: linear-gradient(160deg, rgba(37, 87, 52, 0.12), rgba(77, 51, 32, 0.15));
    border: 1px solid rgba(var(--shadow-color), 0.15);
    border-radius: 22px;
    padding: 1.4rem 1.5rem;
    text-align: left;
    box-shadow: 0 24px 44px rgba(var(--shadow-color), 0.18);
    position: relative;
    overflow: hidden;
    transition: transform 0.22s ease, box-shadow 0.22s ease;
}

.metric-card::after {
    content: "";
    position: absolute;
    top: -50px;
    right: -50px;
    width: 140px;
    height: 140px;
    background: rgba(255, 255, 255, 0.12);
    border-radius: 50%;
}

.metric-card:hover {
    transform: translateY(-4px);
    box-shadow: 0 32px 58px rgba(var(--shadow-color), 0.22);
}

.metric-value {
    font-size: 2.4rem;
    font-weight: 700;
    color: var(--forest-700);
    margin-top: 0.3rem;
}

.metric-label {
    font-size: 0.82rem;
    font-weight: 600;
    letter-spacing: 0.5px;
    color: var(--text-500);
    text-transform: uppercase;
}

.score-table {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    margin-top: 0.6rem;
}

.score-table th,
.score-table td {
    padding: 0.45rem 0.6rem;
    border-bottom: 1px solid rgba(var(--shadow-color), 0.1);
    font-size: 0.9rem;
    text-align: left;
}

.score-table th {
    background: rgba(var(--forest-500), 0.18);
    font-weight: 600;
    color: var(--text-700);
}

.recommendation-chip {
    display: inline-block;
    padding: 4px 14px;
    border-radius: 999px;
    background: linear-gradient(135deg, var(--forest-500), var(--forest-700));
    color: #fefdf8;
    font-size: 0.8rem;
    font-weight: 600;
}

.upload-grid {
    display: grid;
    grid-template-columns: repeat(auto-fit, minmax(240px, 1fr));
    gap: 1.1rem;
    margin: 1.2rem 0;
}

.upload-card {
    background: #ffffff;
    border: 1px dashed rgba(var(--shadow-color), 0.22);
    border-radius: 18px;
    padding: 1.1rem 1.3rem;
    box-shadow: 0 16px 28px rgba(var(--shadow-color), 0.12);
}

.upload-card h4 {
    margin: 0 0 0.55rem 0;
    font-size: 1.02rem;
    color: var(--text-900);
}

.upload-card p {
    margin: 0;
    font-size: 0.86rem;
    color: var(--text-500);
}

div[data-testid="stExpander"] {
    margin-bottom: 1.5rem;
}

div[data-testid="stExpander"] > details {
    border-radius: 22px;
    border: 1px solid rgba(var(--shadow-color), 0.16);
    background: linear-gradient(165deg, rgba(255, 255, 255, 0.98), rgba(241, 234, 223, 0.92));
    box-shadow: 0 26px 52px rgba(var(--shadow-color), 0.18);
    overflow: hidden;
}

div[data-testid="stExpander"] > details > summary {
    font-weight: 700;
    font-size: 1rem;
    color: var(--forest-700);
    padding: 1rem 1.4rem;
    list-style: none;
    position: relative;
}

div[data-testid="stExpander"] > details > summary::before {
    content: "➕";
    margin-right: 0.65rem;
    color: var(--forest-600);
    font-size: 1rem;
}

div[data-testid="stExpander"] > details[open] > summary::before {
    content: "➖";
}

div[data-testid="stExpander"] > details[open] > summary {
    background: rgba(var(--forest-500), 0.14);
    color: var(--forest-800);
}

div[data-testid="stExpander"] > details > div[data-testid="stExpanderContent"] {
    padding: 1.2rem 1.5rem 1.5rem;
    background: #ffffff;
    border-top: 1px solid rgba(var(--shadow-color), 0.12);
}

div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"] {
    border: 1px solid rgba(var(--shadow-color), 0.16);
    border-radius: 22px;
    overflow: hidden;
    box-shadow: 0 22px 44px rgba(var(--shadow-color), 0.18);
    background: #ffffff;
}

div[data-testid="stDataFrame"] div[role="columnheader"],
div[data-testid="stDataEditor"] div[role="columnheader"] {
    background: linear-gradient(120deg, rgba(var(--forest-500), 0.28), rgba(var(--forest-500), 0.18)) !important;
    color: var(--forest-900) !important;
    font-weight: 700;
    font-size: 0.92rem;
    text-transform: uppercase;
    letter-spacing: 0.4px;
    border-bottom: 1px solid rgba(var(--shadow-color), 0.14);
    box-shadow: inset 0 -1px 0 rgba(var(--shadow-color), 0.08);
}

div[data-testid="stDataFrame"] div[role="gridcell"],
div[data-testid="stDataEditor"] div[role="gridcell"] {
    color: var(--text-700);
    font-size: 0.92rem;
    border-bottom: 1px solid rgba(var(--shadow-color), 0.08);
    padding: 0.55rem 0.75rem;
}

div[data-testid="stDataFrame"] div[role="row"],
div[data-testid="stDataEditor"] div[role="row"] {
    transition: background 0.2s ease, box-shadow 0.2s ease;
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div:nth-child(odd) div[role="row"],
div[data-testid="stDataEditor"] div[role="rowgroup"] > div:nth-child(odd) div[role="row"] {
    background: rgba(255, 255, 255, 0.95);
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div:nth-child(even) div[role="row"],
div[data-testid="stDataEditor"] div[role="rowgroup"] > div:nth-child(even) div[role="row"] {
    background: rgba(var(--linen-200), 0.65);
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div div[role="row"]:hover,
div[data-testid="stDataEditor"] div[role="rowgroup"] > div div[role="row"]:hover {
    background: rgba(var(--forest-200), 0.32);
    box-shadow: inset 0 0 0 1px rgba(var(--forest-500), 0.35);
}

div[data-testid="stDataFrame"] div[role="rowgroup"] > div div[role="row"]:hover div[role="gridcell"],
div[data-testid="stDataEditor"] div[role="rowgroup"] > div div[role="row"]:hover div[role="gridcell"] {
    border-bottom-color: transparent;
}
</style>
"""








st.markdown(CSS, unsafe_allow_html=True)


RESULT_COLUMNS = ['evaluacion_numerica', 'sugerencia_rapida']






















def _default_tables():







    return {







        "estatus": pd.DataFrame([







            ("Idea", 12.5), ("Brief", 25.0), ("Modelo", 37.5), ("Prototipo", 50.0),







            ("Conocimiento para futura investigacion", 40.0), ("MVP", 62.5),







            ("Tecnologia", 75.0), ("Servicio", 87.5), ("EBCT", 100.0)







        ], columns=["Concepto", "Valor"]),







        "impacto": pd.DataFrame([







            ("Alto", 30), ("Medio", 20), ("Bajo", 10)







        ], columns=["Concepto", "Valor"]),







        "estado_pm": pd.DataFrame([







            ("Abierto", 10), ("Cerrado", 0)







        ], columns=["Concepto", "Valor"]),







        "activo_pm": pd.DataFrame([







            ("Si", 10), ("No", 0)







        ], columns=["Concepto", "Valor"]),







        "potencial_transferencia": pd.DataFrame([







            ("Bien publico", 10), ("Comercial", 20), ("Uso de transferencia", 30), ("Baja", 0)







        ], columns=["Concepto", "Valor"]),







        "tiene_resp_in": pd.DataFrame([







            ("Si", 0), ("No", 10)







        ], columns=["Concepto", "Valor"]),







        "evaluacion": pd.DataFrame([







            ("Alta", 100), ("Media", 50), ("Baja", 0),







            ("Prioridad_alta_umbral", 375), ("Prioridad_media_umbral", 250)







        ], columns=["Rango", "ValorReferencia"]),







    }























def _sample_portafolio() -> pd.DataFrame:







    return pd.DataFrame([







        {







            "id_innovacion": "P-001", "fecha_creacion": "2024-01-12",







            "nombre_innovacion": "Sensor forestal inteligente", "potencial_transferencia": "Comercial",







            "estatus": "MVP", "impacto": "Alto", "nombre_pm": "Ana Torres", "codigo_pm": "PM-101",







            "responsable_pm": "Ana Torres", "estado_pm": "Abierto", "activo_pm": "Si",







            "responsable_innovacion": "Luis Rojas", "tiene_resp_in": "Si",







            "fecha_inicio_pm": "2024-02-01", "fecha_termino_pm": "2024-09-30",







            "fecha_termino_real_pm": "", "evaluacion_numerica": "320",







            "sugerencia_rapida": "Mantener seguimiento de piloto",







        },







        {







            "id_innovacion": "P-002", "fecha_creacion": "2023-09-03",







            "nombre_innovacion": "Plataforma datos clima", "potencial_transferencia": "Bien publico",







            "estatus": "Servicio", "impacto": "Medio", "nombre_pm": "Carla Mena", "codigo_pm": "PM-089",







            "responsable_pm": "Carla Mena", "estado_pm": "Abierto", "activo_pm": "Si",







            "responsable_innovacion": "Equipo datos", "tiene_resp_in": "No",







            "fecha_inicio_pm": "2023-10-10", "fecha_termino_pm": "2024-08-15",







            "fecha_termino_real_pm": "", "evaluacion_numerica": "260",







            "sugerencia_rapida": "Asignar responsable IN",







        },







        {







            "id_innovacion": "P-003", "fecha_creacion": "2022-05-18",







            "nombre_innovacion": "Modelo prediccion incendios", "potencial_transferencia": "Uso de transferencia",







            "estatus": "EBCT", "impacto": "Alto", "nombre_pm": "Juan Vega", "codigo_pm": "PM-045",







            "responsable_pm": "Juan Vega", "estado_pm": "Abierto", "activo_pm": "Si",







            "responsable_innovacion": "Unidad analitica", "tiene_resp_in": "Si",







            "fecha_inicio_pm": "2022-07-01", "fecha_termino_pm": "2024-12-31",







            "fecha_termino_real_pm": "", "evaluacion_numerica": "410",







            "sugerencia_rapida": "Listo para financiamiento",







        },







        {







            "id_innovacion": "P-004", "fecha_creacion": "2024-03-22",







            "nombre_innovacion": "Manual transferencia", "potencial_transferencia": "Bien publico",







            "estatus": "Modelo", "impacto": "Medio", "nombre_pm": "Marcelo Diaz", "codigo_pm": "PM-120",







            "responsable_pm": "Marcelo Diaz", "estado_pm": "Abierto", "activo_pm": "Si",







            "responsable_innovacion": "Unidad extension", "tiene_resp_in": "Si",







            "fecha_inicio_pm": "2024-04-10", "fecha_termino_pm": "2024-11-30",







            "fecha_termino_real_pm": "", "evaluacion_numerica": "230",







            "sugerencia_rapida": "Revisar contenido legal",







        },







        {







            "id_innovacion": "P-005", "fecha_creacion": "2023-01-09",







            "nombre_innovacion": "App monitoreo viveros", "potencial_transferencia": "Comercial",







            "estatus": "Prototipo", "impacto": "Bajo", "nombre_pm": "Laura Saez", "codigo_pm": "PM-066",







            "responsable_pm": "Laura Saez", "estado_pm": "Cerrado", "activo_pm": "No",







            "responsable_innovacion": "Equipo viveros", "tiene_resp_in": "No",







            "fecha_inicio_pm": "2023-02-01", "fecha_termino_pm": "2023-11-30",







            "fecha_termino_real_pm": "2023-12-15", "evaluacion_numerica": "180",







            "sugerencia_rapida": "Proyecto cerrado por decision externa",







        },







    ])























EXCLUDED_TEMPLATE_COLUMNS = ['evaluacion_numerica', 'sugerencia_rapida']























def _portafolio_template() -> pd.DataFrame:







    base = _sample_portafolio().head(0)







    return base.drop(columns=EXCLUDED_TEMPLATE_COLUMNS, errors='ignore')























def _template_instructions() -> List[str]:







    lines = [







        "Instructivo de Carga - Portafolio Maestro de Innovaciones",







        "",







        "Este instructivo detalla como completar la plantilla de carga masiva.",







        "Los campos evaluacion_numerica y sugerencia_rapida se calculan automaticamente y no van en la plantilla.",







        "",







        "1) Objetivo del archivo",







        "- Registrar innovaciones de manera estandarizada.",







        "- Mantener trazabilidad con el Proyecto Madre (PM).",







        "- Habilitar el calculo automatico de indicadores y priorizacion.",







        "",







        "2) Columnas del archivo (una fila por innovacion)",







        "1. id_innovacion (texto/entero): identificador unico. Ej.: P-101.",







        "2. fecha_creacion (fecha): formato dd-mm-aaaa.",







        "3. nombre_innovacion (texto): titulo claro de la iniciativa.",







        "4. potencial_transferencia (lista): Comercial; Bien publico; Uso de transferencia; Conocimiento para investigacion.",







        "5. estatus (lista): Idea; Brief; Modelo; Prototipo; MVP; Tecnologia; Servicio; EBCT.",







        "6. impacto (lista): Alto; Medio; Bajo.",







        "7. nombre_pm (texto): nombre del proyecto madre.",







        "8. codigo_pm (texto): identificador del PM. Ej.: PM-2025-01.",







        "9. responsable_pm (texto): responsable del PM.",







        "10. estado_pm (lista): Abierto; Cerrado.",







        "11. activo_pm (lista): Si; No.",







        "12. responsable_innovacion (texto): responsable directo de la innovacion.",







        "13. tiene_resp_in (lista): Si; No.",







        "14. fecha_inicio_pm (fecha): formato dd-mm-aaaa.",







        "15. fecha_termino_pm (fecha): formato dd-mm-aaaa.",







        "16. fecha_termino_real_pm (fecha): dejar vacio si sigue en ejecucion.",







        "",







        "Campos calculados (no se incluyen en la plantilla):",







        "- evaluacion_numerica: se crea al ejecutar el Calculo de candidatos.",







        "- sugerencia_rapida: resume alertas y la prioridad resultante.",







        "",







        "3) Listas validas de referencia",







        "- Estatus: Idea; Brief; Modelo; Prototipo; MVP; Tecnologia; Servicio; EBCT.",







        "- Impacto: Alto; Medio; Bajo.",







        "- Estado PM: Abierto; Cerrado.",







        "- Activo PM: Si; No.",







        "- Potencial transferencia: Comercial; Bien publico; Uso de transferencia; Conocimiento para investigacion.",







        "- Tiene Resp IN: Si; No.",







        "",







        "4) Buenas practicas antes de cargar",







        "- Revisar que id_innovacion sea unico.",







        "- Validar que las fechas usen dd-mm-aaaa.",







        "- Confirmar responsables y estados del PM.",







        "- Evitar filas vacias o duplicadas.",







        "",







        "5) Nota",







        "PM = Proyecto Madre. Mantenga consistencia entre nombre_pm y codigo_pm.",







        "",







        "Fin del instructivo.",







    ]







    return lines


























def _build_template_excel(template_df: pd.DataFrame):
    if not HAS_OPENPYXL or Workbook is None or get_column_letter is None:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = 'Plantilla'
    for col_idx, col_name in enumerate(template_df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        if Alignment is not None:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        ws.column_dimensions[get_column_letter(col_idx)].width = max(18, len(col_name) + 4)
    ws.freeze_panes = 'A2'
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()



def _build_instructive_excel(lines: List[str]):
    if not HAS_OPENPYXL or Workbook is None:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = 'Instructivo'
    for idx, line in enumerate(lines, start=1):
        ws.cell(row=idx, column=1, value=line)
    ws.column_dimensions['A'].width = 110
    ws.freeze_panes = 'A2'
    if Alignment is not None:
        for row in ws.iter_rows(min_row=1, max_row=len(lines), max_col=1):
            row[0].alignment = Alignment(wrap_text=True, vertical='top')
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()





def _catalog_options(score_tables: dict) -> dict:
    return {
        'estatus': score_tables['estatus']['Concepto'].tolist(),
        'impacto': score_tables['impacto']['Concepto'].tolist(),
        'estado_pm': score_tables['estado_pm']['Concepto'].tolist(),
        'activo_pm': score_tables['activo_pm']['Concepto'].tolist(),
        'potencial_transferencia': score_tables['potencial_transferencia']['Concepto'].tolist(),
        'tiene_resp_in': score_tables['tiene_resp_in']['Concepto'].tolist(),
    }



def _portafolio_column_config(score_tables: dict) -> dict:
    catalogs = _catalog_options(score_tables)
    config = {}
    for key, options in catalogs.items():
        config[key] = st.column_config.SelectboxColumn(
            label=key.replace('_', ' ').title(),
            options=options,
        )
    return config



def _enforce_catalog_values(df: pd.DataFrame, score_tables: dict):
    catalogs = _catalog_options(score_tables)
    cleaned = df.copy()
    issues = {}
    for key, options in catalogs.items():
        if key not in cleaned.columns:
            continue
        series = cleaned[key].astype(str).fillna('').str.strip()
        mask = (series != '') & ~series.str.lower().isin({str(opt).strip().lower() for opt in options})
        if mask.any():
            issues[key] = sorted(set(series[mask]))
            cleaned.loc[mask, key] = ''
    return cleaned, issues


def _restore_result_columns(df_new: pd.DataFrame, df_original: pd.DataFrame) -> pd.DataFrame:
    df_new = df_new.copy()
    for col in RESULT_COLUMNS:
        if col not in df_new.columns:
            df_new[col] = ''
    existing = [col for col in RESULT_COLUMNS if col in df_original.columns]
    if not existing:
        return df_new
    if 'id_innovacion' in df_new.columns and 'id_innovacion' in df_original.columns:
        lookup = df_original.set_index('id_innovacion')[existing]
        df_new = df_new.set_index('id_innovacion')
        aligned = lookup.reindex(df_new.index)
        for col in existing:
            mask = df_new[col].astype(str).str.strip() == ''
            df_new.loc[mask, col] = aligned.loc[mask, col]
        df_new = df_new.reset_index()
    return df_new


def _prepare_lookup(df: pd.DataFrame):







    col_key, col_val = df.columns[0], df.columns[-1]







    mapping = {}







    for _, row in df.iterrows():







        key = str(row.get(col_key, '')).strip().lower()







        try:







            value = float(row.get(col_val, 0))







        except (TypeError, ValueError):







            value = 0.0







        if key:







            mapping[key] = value







    return mapping























def _thresholds(df_eval: pd.DataFrame):
    lookup = _prepare_lookup(df_eval)
    baja = lookup.get('baja', 0.0)
    media = lookup.get('media', 50.0)
    alta = lookup.get('alta', 100.0)
    if media < baja:
        media = baja
    if alta < media:
        alta = media
    return {
        'baja': baja,
        'media': media,
        'alta': alta,
    }









def _buscar_valor(value, lookup):







    return lookup.get(str(value or '').strip().lower(), 0.0)























def _parse_fecha(value):







    if not value or pd.isna(value):







        return None







    try:







        return pd.to_datetime(value)







    except Exception:







        return None























def calcular_puntaje(row, tablas):







    activo = str(row.get('activo_pm', '')).strip().lower()







    estado = str(row.get('estado_pm', '')).strip().lower()







    if activo == 'no' or estado == 'cerrado':







        return 0.0







    total = 0.0







    total += _buscar_valor(row.get('estatus'), tablas['estatus'])







    total += _buscar_valor(row.get('impacto'), tablas['impacto'])







    total += _buscar_valor(row.get('estado_pm'), tablas['estado_pm'])







    total += _buscar_valor(row.get('potencial_transferencia'), tablas['potencial_transferencia'])







    total += _buscar_valor(row.get('activo_pm'), tablas['activo_pm'])







    total += _buscar_valor(row.get('tiene_resp_in'), tablas['tiene_resp_in'])







    fecha = _parse_fecha(row.get('fecha_termino_pm'))







    if fecha is not None and pd.Timestamp(datetime.now().date()) <= fecha.normalize():







        total += 10.0







    return total























def generar_recomendacion(row, puntaje, tablas):







    partes = []







    estado = str(row.get('estado_pm', '')).strip().lower()







    if estado == 'cerrado':







        partes.append('Proy. cerrado')







    fecha = _parse_fecha(row.get('fecha_termino_pm'))







    if fecha is not None:







        if pd.Timestamp(datetime.now().date()) > fecha.normalize():







            partes.append('Fuera de plazo')







        else:







            partes.append('Dentro de plazo')







    if str(row.get('impacto', '')).strip().lower() == 'alto':







        partes.append('Impacto alto')







    if str(row.get('tiene_resp_in', '')).strip().lower() == 'no':







        partes.append('Sin Resp IN')







    umbrales = _thresholds(tablas['evaluacion'])
    media_lim = umbrales['media']
    alta_lim = umbrales['alta']
    if puntaje <= media_lim:
        partes.append('Prioridad baja')
    elif puntaje <= alta_lim:
        partes.append('Prioridad media')
    else:
        partes.append('Prioridad alta')







    return '; '.join(partes)























fase1_page = next(Path('pages').glob('03_*_Fase_1_TRL.py'), None)















st.title('Fase 0 - Portafolio y filtro inicial')







st.caption('Carga, normaliza y evalua iniciativas antes de avanzar a la radiografia IRL.')















st.divider()















if 'score_tables' not in st.session_state:







    st.session_state['score_tables'] = _default_tables()







score_tables = st.session_state['score_tables']















with st.expander('Configurar tablas de puntaje', expanded=False):







    st.markdown(







        'Puntaje = Estatus + Impacto + Estado PM + Activo PM + Potencial transferencia + Responsable IN + Bono plazo. '







        'El bono es 10 si la fecha de termino declarada sigue vigente.'







    )







    cols_top = st.columns(3)







    cols_bottom = st.columns(3)







    pairs = [







        ('estatus', 'Estatus'),







        ('impacto', 'Impacto'),







        ('estado_pm', 'Estado PM'),







        ('activo_pm', 'Activo PM'),







        ('potencial_transferencia', 'Potencial transferencia'),







        ('tiene_resp_in', 'Tiene Resp IN'),







    ]







    for idx, (key, label) in enumerate(pairs):







        target = cols_top[idx] if idx < 3 else cols_bottom[idx - 3]







        with target:







            st.markdown(f'**{label}**')







            score_tables[key] = st.data_editor(







                score_tables[key],







                num_rows='dynamic',







                hide_index=True,







                use_container_width=True,







                key=f'tabla_{key}',







            )







    st.markdown('**Evaluacion y umbrales**')







    score_tables['evaluacion'] = st.data_editor(







        score_tables['evaluacion'],







        num_rows='dynamic',







        hide_index=True,







        use_container_width=True,







        key='tabla_evaluacion',







    )















st.markdown('<div class="section-card">', unsafe_allow_html=True)







st.markdown('### Editor de portafolio')







st.caption('Agrega o ajusta proyectos antes de ejecutar la estimacion de candidatos.')















portafolio_df = utils.normalize_df(db.fetch_df())







if portafolio_df.empty:







    sample = utils.normalize_df(_sample_portafolio())







    db.replace_all(sample)







    portafolio_df = sample







    st.info('Se cargaron datos de ejemplo para comenzar a trabajar.')















st.markdown('#### Plantillas y carga masiva')







template_df = _portafolio_template()
instructions = _template_instructions()
template_xlsx = _build_template_excel(template_df)
instructivo_xlsx = _build_instructive_excel(instructions)

cols_template = st.columns([1, 1, 2])
with cols_template[0]:
    if template_xlsx is not None:
        st.download_button(
            'Descargar plantilla Excel',
            data=template_xlsx,
            file_name='plantilla_portafolio.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            key='download_template_xlsx',
            use_container_width=True,
        )
    else:
        st.info('Instala openpyxl para habilitar la plantilla de carga en formato Excel.')
with cols_template[1]:
    if instructivo_xlsx is not None:
        st.download_button(
            'Descargar instructivo',
            data=instructivo_xlsx,
            file_name='instructivo_portafolio.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            key='download_instructivo_xlsx',
            use_container_width=True,
        )
    else:
        st.info('Instala openpyxl para descargar el instructivo en Excel.')
with cols_template[2]:
    st.caption(
        'Descarga la plantilla de carga o el instructivo segun necesites y usa la carga masiva para reemplazar o '
        'anexar proyectos.'
    )
uploaded_file = st.file_uploader(







    'Cargar portafolio (CSV o Excel)',







    type=['csv', 'xlsx', 'xls'],







    key='upload_portafolio',







)















if uploaded_file is not None:







    action = st.radio(







        'Como aplicar la carga?',







        ('Reemplazar portafolio actual', 'Anexar al portafolio actual'),







        key='upload_action',







    )







    if not HAS_OPENPYXL and uploaded_file.name.lower().endswith(('xlsx', 'xls')):







        st.error('No es posible leer archivos Excel porque openpyxl no esta instalado.')







    elif st.button('Aplicar carga del archivo', key='btn_aplicar_carga', use_container_width=True):







        try:







            if uploaded_file.name.lower().endswith('.csv'):







                df_import = pd.read_csv(uploaded_file)







            else:







                df_import = pd.read_excel(uploaded_file)







            if df_import.empty:







                st.warning('El archivo no contiene registros.', icon='⚠️')







            else:







                df_import = utils.normalize_df(df_import)







                base_columns = portafolio_df.columns.tolist()







                if not base_columns:







                    base_columns = df_import.columns.tolist()







                additional_cols = [col for col in df_import.columns if col not in base_columns]







                all_columns = list(dict.fromkeys(base_columns + additional_cols))







                if action == 'Anexar al portafolio actual' and not portafolio_df.empty:







                    existing_aligned = portafolio_df.reindex(columns=all_columns)







                    import_aligned = df_import.reindex(columns=all_columns)







                    combined = pd.concat([existing_aligned, import_aligned], ignore_index=True)







                else:







                    combined = df_import.reindex(columns=all_columns)







                df_norm = utils.normalize_df(combined)
                df_norm, invalids = _enforce_catalog_values(df_norm, score_tables)
                df_norm = _restore_result_columns(df_norm, portafolio_df)
                if invalids:
                    details = ' | '.join(f"{col}: {', '.join(vals)}" for col, vals in invalids.items())
                    st.warning(f'Valores fuera de catalogo detectados en la carga: {details}. Se limpiaron para revision.')
                db.replace_all(df_norm)
                portafolio_df = df_norm
                st.session_state.pop('fase0_result', None)
                st.session_state.pop('fase1_payload', None)
                st.session_state.pop('fase1_ready', None)







                st.success('Portafolio actualizado correctamente desde la carga de archivo.')







        except Exception as exc:







            st.error(f'No se pudo procesar el archivo: {exc}')







else:







    st.caption('Selecciona un archivo para activar la carga masiva.')















display_df = portafolio_df.drop(columns=RESULT_COLUMNS, errors='ignore')
with st.expander('Planilla de proyectos (edicion manual)', expanded=False):
    st.caption('Edita la informacion base del portafolio. Los campos de resultado se recalculan cuando vuelves a evaluar.')
    st.markdown('<div class="data-editor">', unsafe_allow_html=True)
    portafolio_editado = st.data_editor(
        display_df,
        num_rows='dynamic',
        hide_index=True,
        use_container_width=True,
        column_config=_portafolio_column_config(score_tables),
        key='editor_portafolio',
    )
    st.markdown('</div>', unsafe_allow_html=True)















if st.button('Guardar portafolio', key='btn_guardar_portafolio'):







    try:







        df_norm = utils.normalize_df(portafolio_editado)
        df_norm, invalids = _enforce_catalog_values(df_norm, score_tables)
        df_norm = _restore_result_columns(df_norm, portafolio_df)
        if invalids:
            details = ' | '.join(f"{col}: {', '.join(vals)}" for col, vals in invalids.items())
            st.warning(f'Valores fuera de catalogo corregidos: {details}. Revisa y ajusta antes de guardar nuevamente.')
        db.replace_all(df_norm)
        portafolio_df = df_norm







        st.session_state.pop('fase0_result', None)
        st.session_state.pop('fase1_payload', None)
        st.session_state.pop('fase1_ready', None)







        st.success('Portafolio actualizado correctamente.')







    except Exception as exc:







        st.error(f'Error al guardar: {exc}')















st.markdown('</div>', unsafe_allow_html=True)







st.divider()















st.markdown('<div class="section-card">', unsafe_allow_html=True)







st.markdown('### Calculo de candidatos')







st.caption('La estimacion usa los puntajes configurados y otorga un bono si la fecha declarada sigue vigente.')















if st.button('Calcular ranking de candidatos', key='btn_calcular'):







    df_eval = utils.normalize_df(db.fetch_df())







    if df_eval.empty:







        st.warning('No hay proyectos para evaluar.')







    else:







        columnas = ['estatus', 'impacto', 'estado_pm', 'activo_pm', 'potencial_transferencia', 'tiene_resp_in', 'fecha_termino_pm']







        for col in columnas:







            if col not in df_eval.columns:







                df_eval[col] = ''







        lookups = {key: _prepare_lookup(score_tables[key]) for key, _ in pairs}







        df_eval['evaluacion_calculada'] = df_eval.apply(lambda row: calcular_puntaje(row, lookups), axis=1)







        df_eval['recomendacion'] = df_eval.apply(







            lambda row: generar_recomendacion(row, row['evaluacion_calculada'], score_tables),







            axis=1,







        )







        df_eval = df_eval.sort_values('evaluacion_calculada', ascending=False).reset_index(drop=True)







        df_eval['ranking'] = np.arange(1, len(df_eval) + 1)







        st.session_state['fase0_result'] = df_eval















resultado = st.session_state.get('fase0_result')







if resultado is not None and not resultado.empty:







    umbrales = _thresholds(score_tables['evaluacion'])







    total = len(resultado)







    candidatos_media = int((resultado['evaluacion_calculada'] > umbrales['media']).sum())















    metric_cards = [
        ('Total proyectos', total),
        ('Candidatos >= prioridad media', candidatos_media),
        ('Puntaje maximo', f"{resultado['evaluacion_calculada'].max():.1f}"),
        ('Puntaje promedio', f"{resultado['evaluacion_calculada'].mean():.1f}"),
    ]
    st.session_state['fase1_payload'] = {
        'ranking': resultado.copy(),
        'metrics_cards': metric_cards.copy(),
        'umbrales': umbrales,
    }
    st.session_state['fase1_ready'] = False
    metric_html = ['<div class="metric-grid">']







    for label, value in metric_cards:







        metric_html.append(







            f'<div class="metric-card"><div class="metric-label">{label}</div><div class="metric-value">{value}</div></div>'







        )







    metric_html.append('</div>')







    st.markdown(''.join(metric_html), unsafe_allow_html=True)















    if fase1_page:







        st.markdown('<div class="primary-btn">', unsafe_allow_html=True)







        if st.button('Ir a Fase 1', key='btn_ir_fase1', type='primary'):
            st.session_state['fase1_ready'] = True
            st.switch_page(str(fase1_page))







        st.markdown('</div>', unsafe_allow_html=True)















    with st.expander('Ranking de candidatos priorizados', expanded=False):
        ranking_display = resultado.copy().reset_index(drop=True)
        if 'evaluacion_calculada' in ranking_display.columns:
            ranking_display['evaluacion_calculada'] = ranking_display['evaluacion_calculada'].astype(float).round(1)

        render_table(
            ranking_display,
            key='fase0_ranking_andes',
            highlight_top_rows=3,
            include_actions=True,
            hide_index=True,
        )

        if HAS_OPENPYXL:
            eval_buffer = BytesIO()

            with pd.ExcelWriter(eval_buffer, engine='openpyxl') as writer:
                resultado.to_excel(writer, index=False, sheet_name='Evaluacion')

                resumen_df = pd.DataFrame([
                    {'Indicador': 'Total proyectos', 'Valor': total},
                    {'Indicador': 'Candidatos >= prioridad media', 'Valor': candidatos_media},
                    {'Indicador': 'Puntaje maximo', 'Valor': f"{resultado['evaluacion_calculada'].max():.1f}"},
                    {'Indicador': 'Puntaje promedio', 'Valor': f"{resultado['evaluacion_calculada'].mean():.1f}"},
                    {'Indicador': 'Umbral prioridad baja', 'Valor': umbrales['baja']},
                    {'Indicador': 'Umbral prioridad media', 'Valor': umbrales['media']},
                    {'Indicador': 'Umbral prioridad alta', 'Valor': umbrales['alta']},
                ])

                resumen_df.to_excel(writer, index=False, sheet_name='Resumen')

            st.download_button(
                'Descargar evaluacion (Excel)',
                data=eval_buffer.getvalue(),
                file_name='evaluacion_fase0.xlsx',
                mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                key='download_eval',
            )
        else:
            st.info('Instala openpyxl para exportar la evaluacion en Excel.')














st.markdown('</div>', unsafe_allow_html=True)


















